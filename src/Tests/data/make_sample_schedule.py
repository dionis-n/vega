#!/usr/bin/env python
"""Генератор тестового .xlsx-расписания для VegaTests.

Формат соответствует тому, что ждёт Parser::loadXLSXFromFile:
  * лист называется "Занятия";
  * день недели ("ПН", "ВТ", ...) — в колонке A;
  * номер пары — в колонке B;
  * для каждой группы блок из 3 колонок: [предмет][.][кабинет],
    первый блок начинается там, где в строке заголовка групп есть текст;
  * строка с названиями групп — на 1 выше первого дня;
  * парсинг прекращается на ячейке со словом "Легенда".

Запуск:  python make_sample_schedule.py  ->  sample_schedule.xlsx
"""
import os
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Занятия"

# Строка 1 — заголовок (игнорируется парсером).
ws["A1"] = "РАСПИСАНИЕ ЗАНЯТИЙ (тестовый файл)"

# Строка 2 — названия групп. Первый блок в колонке C (=> groupColStart = 3),
# второй блок через 3 колонки — в колонке F.
ws["C2"] = "ИКБО-01-24"
ws["F2"] = "ИКБО-02-24"

# Строка 3 и далее — расписание. A=день, B=номер пары,
# группа 1: C=предмет, E=кабинет; группа 2: F=предмет, H=кабинет.
rows = [
    # day,   num, subj1,                     cab1,    subj2,                  cab2
    ("ПН",     1, "лк Математический анализ", "А-101", "лк Математический анализ", "А-101"),
    (None,     2, "Программирование (1пг)",   "Б-202", "Программирование (2пг)",   "Б-203"),
    (None,     3, "2,6,10н Физика",           "В-303", "Iн История",               "Г-404"),
    ("ВТ",     1, "лк Базы данных",           "Д-505", "Экономика",                 "Е-606"),
    (None,     2, "IIн Английский язык",      "Ж-707", None,                        None),
]

r = 3
for day, num, subj1, cab1, subj2, cab2 in rows:
    if day:
        ws.cell(row=r, column=1, value=day)
    ws.cell(row=r, column=2, value=num)
    if subj1:
        ws.cell(row=r, column=3, value=subj1)
        ws.cell(row=r, column=5, value=cab1)
    if subj2:
        ws.cell(row=r, column=6, value=subj2)
        ws.cell(row=r, column=8, value=cab2)
    r += 1

# Маркер конца расписания.
ws.cell(row=r, column=1, value="Легенда")

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_schedule.xlsx")
wb.save(out)
print("Saved:", out)
